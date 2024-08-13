from zipfile import ZipFile
import csv
from pypdf import PdfReader
from openpyxl import load_workbook
from constants import *


def test_archive_csv():
    with ZipFile(ARCHIVE) as zf:
        with zf.open('industry.csv') as csv_file:
            content = csv_file.read().decode()
            csvreader = list(csv.reader(content.splitlines()))
            first_row = csvreader[1]

            assert first_row[0] == 'Accounting/Finance'


def test_archive_pdf():
    with ZipFile(ARCHIVE) as zf:
        with zf.open('Журнал аудита.pdf') as pdf_file:
            content = PdfReader(pdf_file)
            page = content.pages[0]
            text = page.extract_text()

            assert 'Последнее событие' in text


def test_archive_xlsx():
    with ZipFile(ARCHIVE) as zf:
        with zf.open('file_example_XLSX_1000.xlsx') as xlsx_file:
            workbook = load_workbook(xlsx_file)
            sheet = workbook.active
            cell_content = sheet.cell(row=3, column=3).value

            assert cell_content == 'Hashimoto'
